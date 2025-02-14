from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from .forms import SignUpForm, CommentForm, BentoReservationForm, MenuUploadForm, KakeiboForm, SongRequestForm, FavoriteMoviesForm, FavoriteMoviesCommentForm, BBSPostForm, BBSCommentForm
from .models import Post, Comment, BentoReservation, BentoUnavailableDay, User, MenuUpload, KakeiboEntry, SongRequest, FavoriteMovies, FavoriteMoviesComment, BBSPost, BBSComment
from django.urls import resolve, reverse
from .utils import decode_filename
from django.contrib import messages
from datetime import date, datetime, timedelta
from django.utils import timezone
from django.utils.timezone import now
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.template.loader import render_to_string
from django.db.models import Q, Sum, Count
from collections import defaultdict
from django.utils.timezone import localdate
from itertools import groupby
from operator import attrgetter
from django.views.decorators.csrf import csrf_exempt
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from django.conf import settings
import requests
from django.core.cache import cache
from django.urls import reverse

# Create your views here.
def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('index')
    else:
        form = SignUpForm()
    return render(request, 'accounts/signup.html', {'form': form})

@login_required
def index(request):
    posts = Post.objects.all().order_by('-created_at')[:10]
    song_requests = SongRequest.objects.all()
    # 新着情報のデータを取得
    one_day_ago = now() - timedelta(days=1)
    # 新着映画
    new_movies = FavoriteMovies.objects.filter(created_at__gte=one_day_ago)
    # 新着リクエスト
    new_requests = SongRequest.objects.filter(request_date__gte=one_day_ago)
    # 自分の好きな映画への新しいコメント
    recent_favorite_movie_comments = FavoriteMoviesComment.objects.filter(
        created_at__gte=one_day_ago, 
        favorite_movies__user=request.user  # 自分の投稿へのコメント
    ).exclude(user=request.user)  # 自分が投稿したコメントは除外
    # なんでも掲示板
    new_bbs = BBSPost.objects.filter(created_at__gte=one_day_ago).exclude(user=request.user)
    # 自分の掲示板へのコメント
    recent_bbs_comments = BBSComment.objects.filter(
        created_at__gte=one_day_ago,
        post__user=request.user,
        parent_comment=None
    ).exclude(user=request.user)
    # **自分のコメントに対する返信**
    recent_bbs_replies = BBSComment.objects.filter(
        created_at__gte=one_day_ago,
        parent_comment__user=request.user  # 自分のコメントが親コメント
    ).exclude(user=request.user)  # 自分自身の返信は除外
    # 新着情報を1つのリストにまとめる
    recent_updates = []
    for movie in new_movies:
        recent_updates.append({
            "type": "movie",
            "title": f"{movie.user.last_name} {movie.user.first_name}さんが好きな映画を投稿しました",
            "url": reverse("favorite_movies_detail", args=[movie.pk]),
            "created_at": movie.created_at,
        })
    
    for requests in new_requests:
        recent_updates.append({
            "type": "song_request",
            "title": f"{requests.user.last_name} {requests.user.first_name}さんが曲をリクエストしました",
            "url": reverse("song_request_list"),
            "created_at": requests.request_date,
        })
    
    for comment in recent_favorite_movie_comments:
        link = reverse('favorite_movies_detail', args=[comment.favorite_movies.id])
        print(f"Generated link: {link}") 
        recent_updates.append({
            'type': 'comment',
            'title': f"{comment.user.last_name} {comment.user.first_name}さんがあなたの好きな映画にコメントしました",
            'url': link,
            'created_at': comment.created_at,
        })
    
    for bbs in new_bbs:
        recent_updates.append({
            "type": "bbs",
            "title": f"{bbs.user.last_name} {bbs.user.first_name}さんが掲示板に投稿しました",
            "url":reverse("bbs_detail", args=[bbs.pk]),
            "created_at":bbs.created_at,
        })
    
    for bbs_comment in recent_bbs_comments:
        link = reverse('bbs_detail', args=[bbs_comment.post.id])
        print(f"Generated link: {link}") 
        recent_updates.append({
            "type": "bbs_comment",
            "title": f"{bbs_comment.user.last_name} {bbs_comment.user.first_name}さんがあなたの掲示板にコメントしました",
            "url": link,
            "created_at": bbs_comment.created_at,
        })
    
    for reply in recent_bbs_replies:
        recent_updates.append({
            "type": "bbs_reply",
            "title": f"{reply.user.last_name} {reply.user.first_name}さんがあなたのコメントに返信しました",
            "url": reverse('bbs_detail', args=[reply.post.id]),
            "created_at": reply.created_at,
        })

    # 作成日時で並べ替え
    recent_updates.sort(key=lambda x: x["created_at"], reverse=True)
    
    return render(request, 'accounts/index.html', {
        'posts': posts,
        'recent_updates': recent_updates,
    })

def post_list(request):
    posts = Post.objects.all()
    return render(request, 'accounts/post_list.html', {'posts': posts})


@login_required
def post_detail(request, slug):
    post = get_object_or_404(Post, slug=slug)

    # ユーザーがまだ既読でなければ追加
    if request.user not in post.read_by.all():
        post.read_by.add(request.user)


    comments = post.comments.all()
    if request.method == 'POST':
        form = CommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.post = post
            comment.user = request.user
            comment.save()
            return redirect('post_detail', slug=post.slug)
    else:
        form = CommentForm()

    # 新規コメントのフラグを消す
    if post.new_comment:
        post.new_comment = False
        post.save()

    context = {
        'post': post,
        'decoded_file_url': decode_filename(post.attached_file.url) if post.attached_file else None,
        'comments': post.comments.all(),
        'form': form,
    }

    return render(request, 'accounts/post_detail.html', context)

@login_required
def like_post(request, slug):
    post = get_object_or_404(Post, slug=slug)
    if request.user in post.likes.all():
        post.likes.remove(request.user)  # Unlike the post
        liked = False
    else:
        post.likes.add(request.user)  # Like the post
        liked = True
    return JsonResponse({'liked': liked, 'total_likes': post.total_likes()})

@login_required
def add_comment(request, post_id):
    post = get_object_or_404(Post, id=post_id)
    if request.method == 'POST':
        content = request.POST.get('content')
        comment = Comment.objects.create(post=post, user=request.user, content=content)
        post.new_comment = True
        post.save()
    return redirect('post_detail', slug=post.slug)

@login_required
def bento_reservation(request):
    if request.method == 'POST':
        form = BentoReservationForm(request.POST, request=request)  # requestを渡す
        if form.is_valid():
            reservation = form.save(commit=False)
            reservation.user = request.user
            reservation.save()
            messages.success(request, "予約が完了しました。")
            return redirect('reservation_list')
    else:
        form = BentoReservationForm(request=request)  # GETリクエスト時もrequestを渡す


    latest_menu = MenuUpload.objects.last()

    # 献立が存在し、PDFかどうかをチェック
    is_pdf = False
    if latest_menu and latest_menu.file.url.endswith(".pdf"):
        is_pdf = True

    return render(request, 'accounts/bento_reservation.html', {'form': form, 'latest_menu': latest_menu, 'is_pdf': is_pdf})

@login_required
def reservation_list(request):
    # 絞り込み機能：期間を指定してフィルター
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    today = timezone.localdate()  # 現在の日付を取得

    reservations = BentoReservation.objects.filter(user=request.user)

    # 絞り込み条件が指定されている場合
    if start_date:
        reservations = reservations.filter(reservation_date__gte=start_date)
    if end_date:
        reservations = reservations.filter(reservation_date__lte=end_date)

    # 降順で予約を表示
    reservations = reservations.order_by('-reservation_date')

    return render(request, 'accounts/reservation_list.html', {
        'reservations': reservations,
        'start_date': start_date,
        'end_date': end_date,
        'today': today,  # テンプレートに渡す
    })

def cancel_reservation(request, reservation_id):
    reservation = get_object_or_404(BentoReservation, id=reservation_id, user=request.user)

    # 予約日の前日の16時を取得
    cancel_deadline = reservation.reservation_date - timedelta(days=1)
    cancel_deadline = datetime.combine(cancel_deadline, datetime.min.time()).replace(hour=17)

    # 現在の日時が取り消し期限より前かを確認
    if datetime.now() < cancel_deadline:
        reservation.delete()  # 予約を取り消す
        messages.success(request, "予約を取り消しました。")
    else:
        messages.error(request, "予約を取り消せるのは前日の17時までです。")

    return redirect('reservation_list')

def receive_bento(request, reservation_id):
    reservation = get_object_or_404(BentoReservation, id=reservation_id)
    if not reservation.received:
        reservation.received = True
        reservation.save()
    return redirect(reverse('reservation_list'))

def admin_bento_reservation_list(request):
    today = date.today()
    # computed_today: 管理者が画面を見ている日の基準
    computed_today = today

    # computed_next: 本日翌日が予約されていればその日、なければ本日以降で最も早い予約日、なければ本日
    candidate_date = today + timedelta(days=1)
    if BentoReservation.objects.filter(reservation_date=candidate_date).exists():
        computed_next = candidate_date
    else:
        next_reservation = BentoReservation.objects.filter(reservation_date__gt=today).order_by('reservation_date').first()
        computed_next = next_reservation.reservation_date if next_reservation else today

    # GETパラメータがあればそれを使うが、ラベル表示は computed_today, computed_next で行う
    default_start = today.strftime('%Y-%m-%d')
    if 'end_date' in request.GET:
        default_end = request.GET.get('end_date')
    else:
        default_end = computed_next.strftime('%Y-%m-%d')

    start_date_str = request.GET.get('start_date', default_start)
    end_date_str = request.GET.get('end_date', default_end)

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    reservations_by_date = {}
    for n in range((end_date - start_date).days + 1):
        current_date = start_date + timedelta(n)
        reservations = BentoReservation.objects.filter(reservation_date=current_date)
        if reservations.exists():
            reservations_by_date[current_date] = reservations

    side_dish_counts = {}
    rice_100g_counts = {}
    rice_160g_counts = {}
    rice_200g_counts = {}

    if reservations_by_date:
        for current_date, reservations in reservations_by_date.items():
            side_dish_counts[current_date] = reservations.filter(side_dish=True).count()
            rice_100g_counts[current_date] = reservations.filter(rice=True, rice_gram=100).count()
            rice_160g_counts[current_date] = reservations.filter(rice=True, rice_gram=160).count()
            rice_200g_counts[current_date] = reservations.filter(rice=True, rice_gram=200).count()

    return render(request, 'admin/admin_bento_reservation_list.html', {
        'reservations_by_date': reservations_by_date,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'side_dish_counts': side_dish_counts,
        'rice_100g_counts': rice_100g_counts,
        'rice_160g_counts': rice_160g_counts,
        'rice_200g_counts': rice_200g_counts,
        'today_date': computed_today,      # ラベル用：本日
        'next_date': computed_next,          # ラベル用：次回
    })

def export_bento_reservations(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    reservations_by_date = BentoReservation.objects.filter(
        reservation_date__range=[start_date, end_date]
    ).order_by('reservation_date')

    # Excelファイルを作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reservations"

    # ヘッダーのテキストを追加 (1行目)
    header_text = f"{start_date}から{end_date}分 弁当予約者一覧"
    ws.merge_cells('A1:F1')  # ヘッダーテキスト用の結合セル
    header_cell = ws['A1']
    header_cell.value = header_text
    header_cell.font = Font(size=16, bold=True)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ヘッダーを2行目に追加
    headers = ["日付", "氏名", "ごはん", "おかず", "受取済", "振替元"]
    ws.append(headers)

    # ヘッダーのスタイル設定
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # データを3行目から追加
    row_num = 3
    for reservation in reservations_by_date:
        ws.append([
            reservation.reservation_date.strftime("%Y-%m-%d"),
            f"{reservation.user.last_name} {reservation.user.first_name}",
            f"{reservation.rice_gram}g" if reservation.rice else "なし",
            "あり" if reservation.side_dish else "なし",
            "はい" if reservation.received else "いいえ",
            reservation.original_user_name or "なし"
        ])
        row_num += 1

    # 固定幅設定: 日付列の幅を狭くする (例: 15文字幅)
    ws.column_dimensions['A'].width = 15  # 日付列
    # 他の列の自動調整
    for col in ws.iter_cols(min_col=2, max_col=len(headers)):
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 6

    # 枠線を追加
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # フィルターをヘッダー行に追加（2行目にフィルター適用）
    ws.auto_filter.ref = f"A2:F{ws.max_row}"

    # ファイル名に期間を動的に設定
    filename = f"reservations_{start_date}_to_{end_date}.xlsx"

    # レスポンスにExcelファイルを追加
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def create_reservation(request):
    if request.method == 'POST':
        form = BentoReservationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('reservation_list')
        else:
            # フォームが無効な場合、エラーを含んだ状態で再度フォームを表示
            return render(request, 'bento_reservation.html', {'form': form})
    else:
        form = BentoReservationForm()

    return render(request, 'bento_reservation.html', {'form': form})

def generate_order_sheet(request):
    side_dish = request.GET.get('side_dish', 0)
    rice_100g = request.GET.get('rice_100g', 0)
    rice_160g = request.GET.get('rice_160g', 0)
    rice_200g = request.GET.get('rice_200g', 0)
    date = request.GET.get('date', '')

    # デバッグ用のプリント
    print(f"おかず: {side_dish}, 100g: {rice_100g}, 160g: {rice_160g}, 200g: {rice_200g}")

    context = {
        'side_dish': side_dish,
        'rice_100g': rice_100g,
        'rice_160g': rice_160g,
        'rice_200g': rice_200g,
        'date': date,
    }

    return render(request, 'accounts/order_sheet.html', context)

@login_required
def upload_menu(request):
    if request.method == 'POST':
        form = MenuUploadForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('upload_menu')
    else:
        form = MenuUploadForm()

    # 既存のメニューを取得
    menus = MenuUpload.objects.all()

    return render(request, 'accounts/menu.html', {'form': form, 'menus': menus})

@login_required
def delete_menu(request, menu_id):
    menu = get_object_or_404(MenuUpload, id=menu_id)
    file_path = os.path.join(settings.MEDIA_ROOT, str(menu.file))
    menu.delete()
    # ファイルが存在する場合は削除
    if os.path.exists(file_path):
        os.remove(file_path)
    messages.success(request, '献立が削除されました。')
    return redirect('upload_menu')

#家計簿
import logging

logger = logging.getLogger(__name__)

@login_required
def kakeibo_list(request):
    current_year = request.GET.get("year", localdate().year)
    current_month = request.GET.get("month", localdate().month)

    entries = KakeiboEntry.objects.filter(
        user=request.user,
        created_at__year=current_year,
        created_at__month=current_month,
    ).order_by('-created_at')

    # デバッグ用ログ
    logger.info(f"Selected Year: {current_year}, Selected Month: {current_month}")
    logger.info(f"Entries: {entries}")

    grouped_entries = defaultdict(list)
    for entry in entries:
        grouped_entries[entry.created_at].append(entry)

    total_income = entries.filter(transaction_type='income').aggregate(total=Sum('amount'))['total'] or 0
    total_expense = entries.filter(transaction_type='expense').aggregate(total=Sum('amount'))['total'] or 0

    return render(request, 'accounts/kakeibo_list.html', {
        "grouped_entries": dict(grouped_entries),
        "total_income": total_income,
        "total_expense": total_expense,
        "current_year": int(current_year),
        "current_month": int(current_month),
        "month_range": range(1, 13),
        "year_range": range(localdate().year - 5, localdate().year + 1),  # 5年前から今年まで
    })

@login_required
def kakeibo_detail(request, pk):
    entry = get_object_or_404(KakeiboEntry, pk=pk, user=request.user)
    income_categories = ['給与', '副収入', '投資収益', '臨時収入', '不労所得', '返金・補助金', 'その他（収入）']
    expense_categories = ['住居費', '食費', '光熱費', '通信費', '交通費', '保険料', '教育費', '医療費', '娯楽費', '衣服・美容', '交際費', '税金・手数料', 'その他（支出）']

    if request.method == 'POST':
        form = KakeiboForm(request.POST, request.FILES, instance=entry)
        if form.is_valid():
            # 画像削除のチェックが入っている場合
            if 'delete_image' in request.POST and request.POST['delete_image'] == 'on':
                if entry.image:
                    entry.image.delete()  # 画像ファイルを削除
                    entry.image = None  # フィールドを空にする

            form.save()
            messages.success(request, "家計簿が更新されました！")
            return redirect('kakeibo_list')
    else:
        form = KakeiboForm(instance=entry)

    return render(request, 'accounts/kakeibo_detail.html', {
        'entry': entry,
        'income_categories': income_categories,
        'expense_categories': expense_categories,
    })

@login_required
def kakeibo_create(request):
    if request.method == 'POST':
        form = KakeiboForm(request.POST, request.FILES)
        if form.is_valid():
            form.instance.user = request.user  # ユーザー情報を設定
            form.instance.created_at = form.cleaned_data['created_at']
            form.save()  # データベースに保存
            messages.success(request, "収支が追加されました！")
            return redirect('kakeibo_list')  # 保存後にリストページへリダイレクト
        else:
            # フォームが無効な場合、エラーを表示
            messages.error(request, "入力にエラーがあります。確認してください。")
    else:
        form = KakeiboForm()

    return render(request, 'accounts/kakeibo_form.html', {'form': form})

@login_required
def kakeibo_delete(request, pk):
    entry = get_object_or_404(KakeiboEntry, pk=pk, user=request.user)
    entry.delete()
    messages.success(request, "収支データを削除しました！")
    return redirect('kakeibo_list')

#曲リクエスト
@login_required
def song_request_list(request):
    sort = request.GET.get('sort', 'date')  # デフォルトは日付順
    if sort == 'likes':
        requests = SongRequest.objects.annotate(like_count=Count('likes')).order_by('-like_count')
    else:
        requests = SongRequest.objects.order_by('-request_date')
    return render(request, 'accounts/song_request_list.html', {'requests': requests, 'sort': sort})

@login_required
@csrf_exempt  # AJAXリクエスト用
def song_request_create(request):
    if request.method == 'POST':
        data = json.loads(request.body)  # JSONデータを取得
        artist = data.get('artist')
        song_name = data.get('song_name')

        if artist and song_name:
            song_request = SongRequest.objects.create(
                user=request.user,
                artist=artist,
                song_name=song_name
            )
            return JsonResponse({
                'id': song_request.id,
                'artist': song_request.artist,
                'song_name': song_request.song_name,
                'request_date': song_request.request_date.strftime('%Y-%m-%d'),
                'user': f"{song_request.user.last_name} {song_request.user.first_name}",
                'like_count': 0
            })
        return JsonResponse({'error': 'Invalid data'}, status=400)

@login_required
def toggle_like(request, request_id):
    if request.method == 'POST':
        song_request = get_object_or_404(SongRequest, id=request_id)
        if request.user in song_request.likes.all():
            song_request.likes.remove(request.user)
            liked = False
        else:
            song_request.likes.add(request.user)
            liked = True
        song_request.save()
        return JsonResponse({'liked': liked, 'like_count': song_request.likes.count()})
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required
def delete_song_request(request, request_id):
    song_request = get_object_or_404(SongRequest, id=request_id, user=request.user)
    song_request.delete()
    return JsonResponse({'success': True})

#映画
@login_required
def favorite_movies_list(request):
    movies = FavoriteMovies.objects.all().order_by('-created_at')
    return render(request, 'accounts/favorite_movies_list.html', {'movies': movies})

@login_required
def favorite_movies_create(request):
    if request.method == 'POST':
        form = FavoriteMoviesForm(request.POST)
        if form.is_valid():
            favorite_movies = form.save(commit=False)
            favorite_movies.user = request.user
            favorite_movies.save()
            return redirect('favorite_movies_list')
    else:
        form = FavoriteMoviesForm()
    return render(request, 'accounts/favorite_movies_create.html', {'form': form})

@login_required
def favorite_movies_detail(request, pk):
    favorite_movie = get_object_or_404(FavoriteMovies, pk=pk)  # 正しいオブジェクトを取得
    comments = FavoriteMoviesComment.objects.filter(favorite_movies=favorite_movie).order_by('-created_at')

    if request.method == 'POST':
        form = FavoriteMoviesCommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.user = request.user
            comment.favorite_movies = favorite_movie  # 外部キーを正しく設定
            comment.save()
            messages.success(request, "コメントを追加しました。")
            return redirect('favorite_movies_detail', pk=pk)
    else:
        form = FavoriteMoviesCommentForm()

    return render(request, 'accounts/favorite_movies_detail.html', {
        'favorite_movies': favorite_movie,
        'comments': comments,
        'form': form,
    })

@login_required
def favorite_movies_delete(request, pk):
    movie = get_object_or_404(FavoriteMovies, pk=pk)

    # ユーザーが一致する場合のみ削除を許可
    if movie.user == request.user:
        movie.delete()
        return HttpResponseRedirect(reverse('favorite_movies_list'))
    else:
        return HttpResponseRedirect(reverse('favorite_movies_list'))  # 不正な削除試行時のリダイレクト先

#bbs
# なんでも掲示板トップ
def bbs_top(request):
    posts = BBSPost.objects.all().order_by("-created_at")
    return render(request, "accounts/bbs_top.html", {"posts": posts})

# 新規投稿
@login_required
def new_bbs_post(request):
    if request.method == "POST":
        form = BBSPostForm(request.POST)
        if form.is_valid():
            post = form.save(commit=False)
            post.user = request.user
            post.save()
            return redirect("bbs_top")
    else:
        form = BBSPostForm()
    return render(request, "accounts/new_bbs.html", {"form": form})

# 投稿詳細
@login_required
def bbs_detail(request, pk):
    bbs_post = get_object_or_404(BBSPost, pk=pk)
    comment_form = BBSCommentForm()
    reply_form = BBSCommentForm()

    # 親コメントのみ取得（`parent_comment=None` のものだけ）
    comments = BBSComment.objects.filter(post=bbs_post, parent_comment=None).order_by("created_at")

    if request.method == "POST":
        form = BBSCommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.user = request.user
            comment.post = bbs_post
            comment.save()
            return redirect('bbs_detail', pk=pk)

    return render(request, 'accounts/bbs_detail.html', {
        'bbs_post': bbs_post,
        'comments': comments,  # 修正
        'comment_form': comment_form,
        'reply_form': reply_form,
    })

# 投稿削除
@login_required
def delete_bbs_post(request, pk):
    post = get_object_or_404(BBSPost, pk=pk, user=request.user)
    post.delete()
    return redirect("bbs_top")

# コメント投稿
@login_required
def add_bbs_comment(request, pk):
    post = get_object_or_404(BBSPost, pk=pk)
    if request.method == "POST":
        form = BBSCommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.user = request.user
            comment.post = post
            comment.save()
    return redirect("bbs_detail", pk=pk)

# 返信投稿
@login_required
def bbs_reply(request, pk):
    print(f"Received pk: {pk}")  # pk の確認

    # まず、フィルタでオブジェクトが取得できるか確認
    comment_query = BBSComment.objects.filter(pk=pk)
    print(f"Queryset count: {comment_query.count()}")  # 該当オブジェクトの数を確認

    if not comment_query.exists():
        print("Comment not found in filter!")  # もし取得できなければ表示
        return HttpResponse("Comment not found in filter", status=404)

    parent_comment = get_object_or_404(BBSComment, pk=pk)
    print(f"Found parent_comment: {parent_comment.id}")  # ここでエラーになるか確認

    if request.method == "POST":
        content = request.POST.get("content")
        new_reply = BBSComment.objects.create(
            post=parent_comment.post,
            user=request.user,
            content=content,
            parent_comment=parent_comment
        )
        
        
        
        return redirect("bbs_detail", pk=parent_comment.post.pk)